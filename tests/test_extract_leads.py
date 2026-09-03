import tempfile
import unittest
from datetime import datetime
from pathlib import Path

import openpyxl

import extract_leads as engine


class FakeResponse:
    def __init__(self, url, status=200, text="", content_type="text/html", history=None, payload=None):
        self.url = url
        self.status_code = status
        self.text = text
        self.content = text.encode()
        self.headers = {"Content-Type": content_type}
        self.history = history or []
        self._payload = payload or {}

    @property
    def ok(self):
        return 200 <= self.status_code < 400

    def json(self):
        return self._payload


class FakeSession:
    def __init__(self, responses):
        self.responses = responses
        self.headers = {}
        self.calls = []

    def request(self, method, url, **kwargs):
        self.calls.append((method, url))
        return self.responses.get(url, FakeResponse(url, 404, "not found", "text/plain"))

    def post(self, url, **kwargs):
        self.calls.append(("POST", url, kwargs))
        return FakeResponse(url, payload={"places": []})


class ExtractLeadTests(unittest.TestCase):
    def test_email_cleanup_removes_placeholders_and_sentry(self):
        html = "sales@realfirm.com user@domain.com error@sentry.io flags@2x.png"
        self.assertEqual(engine.extract_emails(html), "sales@realfirm.com")
        self.assertFalse(engine.valid_email("user@domain.com"))

    def test_url_normalization_and_internal_links(self):
        normalized = engine.normalize_url("HTTPS://Example.COM/a?utm_source=x&keep=1#section")
        self.assertEqual(normalized, "https://example.com/a?keep=1")
        self.assertTrue(engine.is_internal_url("https://www.example.com/contact", normalized))

    def test_redirect_and_202_are_valid_html_responses(self):
        root = "https://firm.example/"
        final = "https://firm.example/home"
        html = "<html><body><a href='/contact'>Contact</a><p>Services available today.</p></body></html>"
        responses = {
            "https://firm.example/robots.txt": FakeResponse("https://firm.example/robots.txt", 404, "", "text/plain"),
            root: FakeResponse(final, 202, html, history=[object()]),
            "https://firm.example/sitemap.xml": FakeResponse("https://firm.example/sitemap.xml", 404, "", "text/plain"),
            "https://firm.example/contact": FakeResponse("https://firm.example/contact", 200, html),
        }
        issues = []
        prospect = engine.Prospect(company_name="Firm", website_original=root)
        crawl = engine.WebCrawler(engine.Config(max_pages_per_site=2), issues, FakeSession(responses)).validate_and_crawl(prospect)
        self.assertTrue(crawl.website_reachable)
        self.assertEqual(crawl.final_http_status, "202")
        self.assertEqual(crawl.redirect_count, 1)
        self.assertEqual(crawl.crawl_status, "COMPLETED")

    def test_jsonld_chat_booking_and_internal_page_detection(self):
        html = """<html><body><h1>Roof Repair</h1><a href='/contact'>Contact</a>
        <a href='https://calendly.com/firm/intro'>Book now</a>
        <script type='application/ld+json'>{"@type":"LocalBusiness","name":"Roof Firm",
        "telephone":"512-555-1212","email":"sales@roof.example","service":"Roof repair",
        "areaServed":"Austin"}</script>
        <script src='https://widget.intercom.io/widget.js'></script></body></html>"""
        page = engine.PageRecord("https://roof.example/", 200, html, True, 0)
        data = engine.extract_site_data(engine.Prospect(company_name="Roof"), engine.CrawlResult(
            original_url=page.url, final_url=page.url, website_reachable=True, pages=[page]
        ))
        self.assertTrue(data.jsonld_found)
        self.assertTrue(data.has_chat)
        self.assertEqual(data.chat_provider, "intercom")
        self.assertTrue(data.has_booking)
        self.assertEqual(data.booking_provider, "calendly")
        self.assertEqual(data.verified_email, "sales@roof.example")
        self.assertEqual(data.verified_phone, "+1 (512) 555-1212")

    def test_score_is_bounded_and_tiered(self):
        config = engine.Config()
        data = engine.ExtractedData(high_ticket_likelihood="TRUE", verified_phone="+1 (512) 555-1212")
        prospect = engine.Prospect(google_rating=5.0, google_review_count=100, target_category="Home Services")
        score = engine.calculate_ai_fit_score(prospect, data, config)
        self.assertGreaterEqual(score, 0)
        self.assertLessEqual(score, 100)
        self.assertEqual(engine.opportunity_tier(87, config), "A+")
        self.assertEqual(engine.opportunity_tier(70, config), "A")
        self.assertEqual(engine.opportunity_tier(39, config), "D")

    def test_deduplication_retains_search_queries(self):
        first = engine.Prospect(company_name="A", website_original="https://example.com", search_query="one")
        second = engine.Prospect(company_name="A Inc", website_original="https://example.com", search_query="two")
        merged = engine.deduplicate_prospects([first, second])
        self.assertEqual(len(merged), 1)
        self.assertIn("one", merged[0].search_query)
        self.assertIn("two", merged[0].search_query)

    def test_unique_filename_and_workbook_sheets(self):
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory)
            timestamp = datetime(2026, 9, 3, 14, 30, 25)
            first = engine.unique_output_path(output_dir, timestamp)
            first.touch()
            second = engine.unique_output_path(output_dir, timestamp)
            self.assertNotEqual(first, second)
            prospect = engine.Prospect(company_name="Fixture", ai_fit_score=87, opportunity_tier="A+")
            output = engine.export_workbook([prospect], [], {"run_id": "test"}, second)
            workbook = openpyxl.load_workbook(output)
            self.assertEqual(set(workbook.sheetnames), {
                "Prospects", "Top Opportunities", "Demo Candidates", "Validation Issues", "Run Summary"
            })
            self.assertEqual(workbook["Prospects"].max_row, 2)
            self.assertTrue(workbook["Prospects"]["A1"].font.color.rgb.endswith("FFFFFF"))


if __name__ == "__main__":
    unittest.main()
