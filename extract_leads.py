#!/usr/bin/env python3
"""FORM4TH B2B Prospect Engine.

Flujo: discovery -> deduplication -> validation -> controlled crawl ->
extraction -> deterministic AI Front Desk fit scoring -> timestamped XLSX.

The engine does not require an AI API. All observations are based on data
returned by Google Places API (New) and publicly reachable HTML pages.
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import logging
import os
import re
import subprocess
import sys
import time
from collections import deque
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse
from urllib.robotparser import RobotFileParser


# ---------------------------------------------------------------------------
# Auto-installer: dependency checks happen before third-party imports.
# ---------------------------------------------------------------------------
REQUIRED_PACKAGES = {
    "pandas": "pandas",
    "openpyxl": "openpyxl",
    "bs4": "beautifulsoup4",
    "requests": "requests",
}


def ensure_dependencies() -> None:
    """Install only missing packages using the active Python interpreter."""
    missing = [
        package
        for import_name, package in REQUIRED_PACKAGES.items()
        if importlib.util.find_spec(import_name) is None
    ]
    if not missing:
        return
    print(f"[SETUP] Installing dependencies: {', '.join(missing)}")
    try:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--disable-pip-version-check", *missing]
        )
    except subprocess.CalledProcessError as exc:
        raise SystemExit(
            "Could not install dependencies. Run: "
            f"{sys.executable} -m pip install {' '.join(missing)}"
        ) from exc


ensure_dependencies()

import pandas as pd  # noqa: E402
import requests  # noqa: E402
from bs4 import BeautifulSoup  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402


# ---------------------------------------------------------------------------
# Configuration and constants
# ---------------------------------------------------------------------------
PLACES_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"
GOOGLE_PLACES_API_KEY = os.getenv("GOOGLE_PLACES_API_KEY", "")
DEFAULT_OUTPUT_DIR = Path("output")
DEFAULT_USER_AGENT = "FORM4THProspectEngine/1.0 (+public-business-research)"
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
PHONE_RE = re.compile(r"(?:\+?1[\s.\-]?)?(?:\(?\d{3}\)?[\s.\-])\d{3}[\s.\-]\d{4}")
ASSET_EXTENSIONS = {
    ".7z", ".avi", ".bmp", ".css", ".csv", ".doc", ".docx", ".eot", ".gif",
    ".ico", ".jpg", ".jpeg", ".js", ".json", ".mp3", ".mp4", ".png", ".svg",
    ".tar", ".ttf", ".webm", ".webp", ".woff", ".woff2", ".xls", ".xlsx", ".zip",
}
PRIORITY_TERMS = (
    "services", "service", "solutions", "products", "about", "contact", "faq",
    "pricing", "plans", "locations", "areas-we-serve", "schedule", "booking",
    "appointment", "consultation", "quote", "estimate", "offers", "financing",
    "team", "properties",
)
EXCLUDED_EMAIL_LOCAL_PARTS = {
    "user", "test", "tester", "example", "sample", "email", "yourname", "name",
    "noreply", "no-reply", "donotreply", "do-not-reply",
}
EXCLUDED_EMAIL_DOMAINS = {
    "domain.com", "example.com", "sentry.io", "sentry.wixpress.com",
    "wixpress.com", "localhost", "invalid",
}
CHAT_PROVIDERS = {
    "intercom": ("intercom",),
    "hubspot": ("hubspot.com", "hs-script-loader", "hbspt"),
    "drift": ("drift.com", "driftt"),
    "tidio": ("tidio.co", "tidiochat"),
    "zendesk": ("zendesk", "zopim"),
    "livechat": ("livechatinc", "livechat.com"),
    "podium": ("podium.com", "podium"),
    "birdeye": ("birdeye.com", "birdeye"),
    "gohighlevel": ("leadconnectorhq", "gohighlevel", "highlevel"),
    "crisp": ("crisp.chat", "crisp-client"),
    "tawk.to": ("tawk.to", "tawk_"),
}
BOOKING_PROVIDERS = {
    "calendly": ("calendly.com", "calendly"),
    "google_calendar": ("calendar.google.com", "calendar.app.google", "google calendar"),
    "acuity": ("acuityscheduling.com", "acuity"),
    "hubspot_meetings": ("meetings.hubspot.com", "hubspot meetings"),
    "gohighlevel": ("leadconnectorhq", "gohighlevel", "highlevel"),
    "schedule_engine": ("scheduleengine.com", "schedule engine"),
    "housecall_pro": ("housecallpro.com", "housecall pro"),
    "servicetitan": ("servicetitan.com", "service titan"),
}
FIELD_MASK = (
    "places.id,places.displayName,places.formattedAddress,places.nationalPhoneNumber,"
    "places.websiteUri,places.rating,places.userRatingCount,places.primaryTypeDisplayName"
)
DEFAULT_QUERIES = {
    "Home Services": [
        "Roofing contractors in Austin TX", "HVAC services in Phoenix AZ",
    ],
    "Law Firms": [
        "Personal injury lawyer in Dallas TX", "Family law attorney in Houston TX",
    ],
    "Real Estate": [
        "Property management in Miami FL", "Commercial real estate in Atlanta GA",
    ],
}
CATEGORY_ALIASES = {
    "home_services": "Home Services",
    "home-services": "Home Services",
    "law_firms": "Law Firms",
    "law-firms": "Law Firms",
    "real_estate": "Real Estate",
    "real-estate": "Real Estate",
}


@dataclass
class Config:
    """Centralized runtime, crawl and scoring configuration."""

    max_pages_per_site: int = 12
    max_depth: int = 1
    request_timeout: int = 5
    discovery_timeout: int = 15
    retries: int = 1
    rate_limit_seconds: float = 0.25
    page_size: int = 20
    output_dir: Path = DEFAULT_OUTPUT_DIR
    min_reviews: int = 0
    weights: dict[str, int] = field(default_factory=lambda: {
        "high_ticket_industry": 25,
        "no_ai_chat": 20,
        "no_easy_booking": 10,
        "after_hours_gap": 10,
        "easy_contact": 15,
        "public_demo_data": 10,
        "strong_business_signals": 10,
    })
    penalties: dict[str, int] = field(default_factory=lambda: {
        "effective_booking": 10,
        "sophisticated_chatbot": 20,
        "strong_24_7_intake": 10,
        "sophisticated_funnel": 15,
        "healthcare_complexity": 15,
        "legal_complexity": 10,
        "large_enterprise": 10,
    })
    tier_thresholds: dict[str, int] = field(default_factory=lambda: {
        "A+": 85, "A": 70, "B": 55, "C": 40,
    })


@dataclass
class ValidationIssue:
    company_name: str
    url: str
    issue_type: str
    http_status: str
    error: str
    stage: str
    timestamp: str


@dataclass
class PageRecord:
    url: str
    status: int
    html: str
    is_html: bool
    depth: int
    error: str = ""


@dataclass
class CrawlResult:
    original_url: str
    final_url: str = ""
    homepage_status: str = ""
    final_http_status: str = ""
    redirect_count: int = 0
    website_reachable: bool = False
    internal_link_count: int = 0
    estimated_page_count: int | str = "UNKNOWN"
    sitemap_found: bool = False
    robots_txt_found: bool = False
    javascript_heavy: bool = False
    crawl_status: str = "NOT_STARTED"
    crawl_error: str = ""
    contact_page: str = "UNKNOWN"
    pages: list[PageRecord] = field(default_factory=list)


@dataclass
class EmailEvidence:
    email: str
    source_url: str
    source_type: str
    confidence: str


@dataclass
class ExtractedData:
    email_candidates: str = "N/A"
    verified_email: str = "N/A"
    email_confidence: str = "NONE"
    email_source_url: str = "UNKNOWN"
    verified_phone: str = "N/A"
    phone_source_url: str = "UNKNOWN"
    hours_source_url: str = "UNKNOWN"
    booking_source_url: str = "UNKNOWN"
    services_source_urls: str = "UNKNOWN"
    jsonld_found: bool = False
    schema_types: str = "UNKNOWN"
    business_name: str = "UNKNOWN"
    industry: str = "UNKNOWN"
    services: str = "UNKNOWN"
    service_area: str = "UNKNOWN"
    locations: str = "UNKNOWN"
    office_hours: str = "UNKNOWN"
    faq_detected: bool = False
    has_chat: bool = False
    chat_provider: str = "UNKNOWN"
    has_ai_chat: bool = False
    has_booking: bool = False
    booking_provider: str = "UNKNOWN"
    booking_url: str = "UNKNOWN"
    has_contact_form: bool = False
    has_quote_form: bool = False
    has_lead_capture: bool = False
    limited_business_hours: str = "UNKNOWN"
    after_hours_gap: str = "UNKNOWN"
    main_cta: str = "UNKNOWN"
    main_conversion_goal: str = "UNKNOWN"
    business_model: str = "UNKNOWN"
    high_ticket_likelihood: str = "UNKNOWN"
    business_sophistication: str = "UNKNOWN"
    evidence_urls: str = "UNKNOWN"


@dataclass
class Prospect:
    """Flat export model; original discovery data is never discarded."""

    company_name: str = "UNKNOWN"
    target_category: str = "UNKNOWN"
    search_query: str = "UNKNOWN"
    primary_type: str = "UNKNOWN"
    address: str = "UNKNOWN"
    city: str = "UNKNOWN"
    state: str = "UNKNOWN"
    phone_original: str = "N/A"
    email_original: str = "N/A"
    website_original: str = "N/A"
    # Convenience fields for a future demo builder; originals remain intact.
    phone: str = "N/A"
    email: str = "N/A"
    google_rating: float | str = "UNKNOWN"
    google_review_count: int = 0
    verified_website: str = "N/A"
    verified_phone: str = "N/A"
    verified_email: str = "N/A"
    contact_page: str = "UNKNOWN"
    website_reachable: bool = False
    final_url: str = "UNKNOWN"
    final_http_status: str = "UNKNOWN"
    redirect_count: int = 0
    validation_status: str = "UNVERIFIED"
    validation_notes: str = "UNKNOWN"
    last_verified_at: str = "UNKNOWN"
    homepage_status: str = "UNKNOWN"
    internal_link_count: int = 0
    estimated_page_count: int | str = "UNKNOWN"
    sitemap_found: bool = False
    robots_txt_found: bool = False
    javascript_heavy: bool = False
    crawl_status: str = "NOT_STARTED"
    crawl_error: str = ""
    ai_fit_score: int = 0
    opportunity_tier: str = "D"
    recommended_action: str = "SKIP"
    sales_angle: str = "UNKNOWN"
    sales_angle_evidence: str = "UNKNOWN"
    qualification_flow_recommendation: str = "manual_review"
    suggested_first_question: str = "UNKNOWN"
    suggested_conversion_goal: str = "UNKNOWN"
    phone_source_url: str = "UNKNOWN"
    email_source_url: str = "UNKNOWN"
    hours_source_url: str = "UNKNOWN"
    booking_source_url: str = "UNKNOWN"
    services_source_urls: str = "UNKNOWN"
    email_candidates: str = "N/A"
    email_confidence: str = "NONE"
    jsonld_found: bool = False
    schema_types: str = "UNKNOWN"
    business_name: str = "UNKNOWN"
    industry: str = "UNKNOWN"
    services: str = "UNKNOWN"
    service_area: str = "UNKNOWN"
    locations: str = "UNKNOWN"
    office_hours: str = "UNKNOWN"
    faq_detected: bool = False
    has_chat: bool = False
    chat_provider: str = "UNKNOWN"
    has_ai_chat: bool = False
    has_booking: bool = False
    booking_provider: str = "UNKNOWN"
    booking_url: str = "UNKNOWN"
    has_contact_form: bool = False
    has_quote_form: bool = False
    has_lead_capture: bool = False
    limited_business_hours: str = "UNKNOWN"
    after_hours_gap: str = "UNKNOWN"
    main_cta: str = "UNKNOWN"
    main_conversion_goal: str = "UNKNOWN"
    business_model: str = "UNKNOWN"
    high_ticket_likelihood: str = "UNKNOWN"
    business_sophistication: str = "UNKNOWN"
    evidence_urls: str = "UNKNOWN"

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


# ---------------------------------------------------------------------------
# Utility, discovery and deduplication
# ---------------------------------------------------------------------------
def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def clean_value(value: Any, default: str = "UNKNOWN") -> str:
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip()


def normalize_phone(phone: str | None) -> str:
    digits = re.sub(r"\D", "", phone or "")
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    return digits


def format_us_phone(phone: str | None) -> str:
    digits = normalize_phone(phone)
    if len(digits) == 10:
        return f"+1 ({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return f"+{digits}" if digits else "N/A"


def normalize_url(url: str | None, base_url: str | None = None) -> str:
    """Normalize URL, remove fragments/tracking and keep only web schemes."""
    if not url:
        return ""
    candidate = urljoin(base_url or "", url.strip())
    parsed = urlparse(candidate)
    if parsed.scheme.lower() not in {"http", "https"} or not parsed.netloc:
        return ""
    filtered_query = [
        (key, value)
        for key, value in parse_qsl(parsed.query, keep_blank_values=True)
        if not key.lower().startswith(("utm_", "gclid", "fbclid", "mc_"))
    ]
    return urlunparse((
        parsed.scheme.lower(), parsed.netloc.lower(), parsed.path or "/",
        "", urlencode(filtered_query), "",
    ))


def registered_domain(url: str) -> str:
    host = (urlparse(url).hostname or "").lower().removeprefix("www.")
    # This intentionally avoids a public-suffix dependency. It is sufficient
    # for the US business domains targeted by this project.
    return host


def is_asset_url(url: str) -> bool:
    return Path(urlparse(url).path.lower()).suffix in ASSET_EXTENSIONS


def is_internal_url(url: str, root_url: str) -> bool:
    return registered_domain(url) == registered_domain(root_url)


def safe_text(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(safe_text(item) for item in value if item)
    if isinstance(value, dict):
        if value.get("name") or value.get("text") or value.get("@id"):
            return clean_value(value.get("name") or value.get("text") or value.get("@id"))
        address_parts = [value.get(key) for key in ("streetAddress", "addressLocality", "addressRegion", "postalCode")]
        return ", ".join(str(item).strip() for item in address_parts if item) or "UNKNOWN"
    return clean_value(value)


def parse_city_state(address: str) -> tuple[str, str]:
    parts = [part.strip() for part in address.split(",") if part.strip()]
    for index, part in enumerate(parts):
        match = re.match(r"^([A-Z]{2})(?:\s+\d{5}(?:-\d{4})?)?$", part)
        if match and index > 0:
            return parts[index - 1], match.group(1)
    return "UNKNOWN", "UNKNOWN"


def extract_display_name(place: dict[str, Any]) -> str:
    value = place.get("displayName", "")
    return clean_value(value.get("text") if isinstance(value, dict) else value)


def build_place_prospect(place: dict[str, Any], category: str, query: str) -> Prospect:
    address = clean_value(place.get("formattedAddress"), "UNKNOWN")
    city, state = parse_city_state(address)
    primary_type = place.get("primaryTypeDisplayName", "")
    if isinstance(primary_type, dict):
        primary_type = primary_type.get("text", "")
    return Prospect(
        company_name=extract_display_name(place),
        target_category=category,
        search_query=query,
        primary_type=clean_value(primary_type),
        address=address,
        city=city,
        state=state,
        phone_original=clean_value(place.get("nationalPhoneNumber"), "N/A"),
        website_original=clean_value(place.get("websiteUri"), "N/A"),
        google_rating=place.get("rating", "UNKNOWN"),
        google_review_count=int(place.get("userRatingCount") or 0),
    )


class PlacesNewDiscovery:
    """Official REST discovery client for Google Places API (New)."""

    def __init__(self, api_key: str, config: Config, session: requests.Session | None = None):
        self.api_key = api_key
        self.config = config
        self.session = session or requests.Session()

    def search(self, query: str) -> list[dict[str, Any]]:
        headers = {
            "Content-Type": "application/json",
            "X-Goog-Api-Key": self.api_key,
            "X-Goog-FieldMask": FIELD_MASK,
        }
        body = {"textQuery": query, "pageSize": self.config.page_size}
        response = self.session.post(
            PLACES_SEARCH_URL,
            headers=headers,
            json=body,
            timeout=self.config.discovery_timeout,
        )
        if not response.ok:
            raise RuntimeError(
                f"Google Places API HTTP {response.status_code}: {response.text[:500]}"
            )
        payload = response.json()
        places = payload.get("places", [])
        return places if isinstance(places, list) else []


def dedupe_key(prospect: Prospect) -> str:
    domain = registered_domain(prospect.website_original) if prospect.website_original != "N/A" else ""
    phone = normalize_phone(prospect.phone_original)
    name = re.sub(r"[^a-z0-9]", "", prospect.company_name.lower())
    address = re.sub(r"[^a-z0-9]", "", prospect.address.lower())
    if domain:
        return f"domain:{domain}"
    if phone:
        return f"phone:{phone}"
    return f"name_address:{name}:{address}"


def deduplicate_prospects(prospects: Iterable[Prospect]) -> list[Prospect]:
    """Consolidate duplicate businesses while retaining all source queries."""
    merged: dict[str, Prospect] = {}
    for prospect in prospects:
        key = dedupe_key(prospect)
        if key not in merged:
            merged[key] = prospect
            continue
        current = merged[key]
        queries = {item.strip() for item in current.search_query.split(" | ") if item.strip()}
        queries.add(prospect.search_query)
        current.search_query = " | ".join(sorted(queries))
        if current.website_original == "N/A" and prospect.website_original != "N/A":
            current.website_original = prospect.website_original
        if current.phone_original == "N/A" and prospect.phone_original != "N/A":
            current.phone_original = prospect.phone_original
        if current.google_review_count < prospect.google_review_count:
            current.google_review_count = prospect.google_review_count
            current.google_rating = prospect.google_rating
    return list(merged.values())


# ---------------------------------------------------------------------------
# Validation and controlled crawl
# ---------------------------------------------------------------------------
class WebCrawler:
    def __init__(self, config: Config, issues: list[ValidationIssue], session: requests.Session | None = None):
        self.config = config
        self.issues = issues
        self.session = session or requests.Session()
        self.session.headers.update({"User-Agent": DEFAULT_USER_AGENT})
        self.last_request_at: dict[str, float] = {}

    def _rate_limit(self, url: str) -> None:
        domain = registered_domain(url)
        previous = self.last_request_at.get(domain, 0)
        wait = self.config.rate_limit_seconds - (time.monotonic() - previous)
        if wait > 0:
            time.sleep(wait)
        self.last_request_at[domain] = time.monotonic()

    def request(self, url: str, method: str = "GET") -> requests.Response:
        last_error: Exception | None = None
        for attempt in range(self.config.retries + 1):
            self._rate_limit(url)
            try:
                response = self.session.request(
                    method, url, timeout=self.config.request_timeout, allow_redirects=True
                )
                if response.status_code >= 500 and attempt < self.config.retries:
                    time.sleep(0.25 * (2**attempt))
                    continue
                return response
            except requests.RequestException as exc:
                last_error = exc
                if attempt < self.config.retries:
                    time.sleep(0.25 * (2**attempt))
        raise last_error or RuntimeError("request failed")

    def _robots(self, root_url: str) -> tuple[RobotFileParser | None, bool, list[str]]:
        robots_url = urljoin(root_url, "/robots.txt")
        try:
            response = self.request(robots_url)
            if response.status_code == 200:
                parser = RobotFileParser()
                parser.set_url(robots_url)
                parser.parse(response.text.splitlines())
                sitemaps = [
                    line.split(":", 1)[1].strip()
                    for line in response.text.splitlines()
                    if line.lower().startswith("sitemap:") and ":" in line
                ]
                return parser, True, sitemaps
        except requests.RequestException:
            pass
        return None, False, []

    def _sitemap(self, root_url: str, robots_sitemaps: list[str]) -> tuple[bool, int | str]:
        candidates = robots_sitemaps or [urljoin(root_url, "/sitemap.xml")]
        for sitemap_url in candidates[:2]:
            try:
                response = self.request(sitemap_url)
                if response.status_code == 200 and "<url" in response.text.lower():
                    count = len(re.findall(r"<loc(?:\s[^>]*)?>(.*?)</loc>", response.text, re.I | re.S))
                    return True, count or "UNKNOWN"
            except requests.RequestException:
                continue
        return False, "UNKNOWN"

    def validate_and_crawl(self, prospect: Prospect) -> CrawlResult:
        result = CrawlResult(original_url=prospect.website_original)
        if prospect.website_original in {"N/A", "UNKNOWN", ""}:
            result.crawl_status = "NO_WEBSITE"
            result.crawl_error = "No website supplied by Places API"
            prospect.validation_status = "UNVERIFIED"
            prospect.validation_notes = result.crawl_error
            prospect.last_verified_at = now_iso()
            return result

        original = normalize_url(prospect.website_original)
        if not original:
            result.crawl_status = "INVALID_URL"
            result.crawl_error = "Unsupported or malformed URL"
            prospect.validation_status = "INVALID"
            prospect.validation_notes = result.crawl_error
            prospect.last_verified_at = now_iso()
            self._issue(prospect, original, "invalid_url", "", result.crawl_error, "validation")
            return result

        try:
            parser, robots_found, robots_sitemaps = self._robots(original)
            result.robots_txt_found = robots_found
            if parser and not parser.can_fetch(DEFAULT_USER_AGENT, original):
                result.crawl_status = "BLOCKED_BY_ROBOTS"
                result.crawl_error = "Homepage disallowed by robots.txt"
                prospect.validation_status = "ERROR"
                prospect.validation_notes = result.crawl_error
                prospect.last_verified_at = now_iso()
                self._issue(prospect, original, "robots_restriction", "", result.crawl_error, "robots")
                return result

            homepage = self.request(original)
            result.final_url = normalize_url(homepage.url) or original
            result.final_http_status = str(homepage.status_code)
            result.redirect_count = len(homepage.history)
            result.homepage_status = str(homepage.status_code)
            content_type = homepage.headers.get("Content-Type", "").lower()
            html = homepage.text if ("html" in content_type or "<html" in homepage.text[:1000].lower()) else ""
            result.website_reachable = 200 <= homepage.status_code < 400 and bool(html)
            if not result.website_reachable:
                result.crawl_status = "INVALID_RESPONSE"
                result.crawl_error = "Response was not a reachable HTML page"
                prospect.validation_status = "INVALID"
                self._issue(prospect, result.final_url, "inaccessible_page", result.final_http_status, result.crawl_error, "validation")
                return result

            soup = BeautifulSoup(html, "html.parser")
            visible_words = len(soup.get_text(" ", strip=True).split())
            result.javascript_heavy = visible_words < 80 and len(soup.find_all("script")) >= 3
            result.pages.append(PageRecord(result.final_url, homepage.status_code, html, True, 0))
            parser_root = result.final_url or original
            result.sitemap_found, result.estimated_page_count = self._sitemap(parser_root, robots_sitemaps)
            self._crawl_pages(prospect, result, parser, parser_root)
            result.internal_link_count = len({
                url for page in result.pages for url in self._links(page.html, page.url) if is_internal_url(url, parser_root)
            })
            if result.estimated_page_count == "UNKNOWN":
                result.estimated_page_count = max(1, result.internal_link_count + 1)
            result.crawl_status = "COMPLETED"
            prospect.validation_status = "PARTIALLY_VERIFIED" if result.javascript_heavy else "VERIFIED"
            prospect.validation_notes = "JavaScript-heavy homepage; extracted server HTML may be incomplete" if result.javascript_heavy else "Reachable HTML verified"
        except requests.Timeout as exc:
            result.crawl_status, result.crawl_error = "ERROR", "Timeout"
            self._issue(prospect, original, "timeout", "", str(exc), "validation")
            prospect.validation_status = "ERROR"
        except requests.RequestException as exc:
            result.crawl_status, result.crawl_error = "ERROR", type(exc).__name__
            self._issue(prospect, original, "inaccessible_page", "", str(exc), "validation")
            prospect.validation_status = "ERROR"
        except Exception as exc:
            result.crawl_status, result.crawl_error = "ERROR", type(exc).__name__
            self._issue(prospect, original, "malformed_html", result.final_http_status, str(exc), "crawl")
            prospect.validation_status = "ERROR"
        prospect.last_verified_at = now_iso()
        return result

    def _issue(self, prospect: Prospect, url: str, issue_type: str, status: str, error: str, stage: str) -> None:
        self.issues.append(ValidationIssue(
            prospect.company_name, url, issue_type, status, error, stage, now_iso()
        ))

    def _links(self, html: str, page_url: str) -> list[str]:
        soup = BeautifulSoup(html, "html.parser")
        links = []
        for anchor in soup.find_all("a", href=True):
            link = normalize_url(anchor.get("href"), page_url)
            if link and is_internal_url(link, page_url) and not is_asset_url(link):
                links.append(link)
        return list(dict.fromkeys(links))

    def _crawl_pages(self, prospect: Prospect, result: CrawlResult, parser: RobotFileParser | None, root_url: str) -> None:
        seen = {result.final_url}
        queue: deque[tuple[str, int]] = deque()
        homepage_links = self._links(result.pages[0].html, root_url)
        homepage_links.sort(key=lambda url: self._priority(url))
        queue.extend((link, 1) for link in homepage_links)
        while queue and len(result.pages) < self.config.max_pages_per_site:
            url, depth = queue.popleft()
            if url in seen or depth > self.config.max_depth:
                continue
            seen.add(url)
            if parser and not parser.can_fetch(DEFAULT_USER_AGENT, url):
                continue
            try:
                response = self.request(url)
                content_type = response.headers.get("Content-Type", "").lower()
                html = response.text if "html" in content_type or "<html" in response.text[:1000].lower() else ""
                if not html:
                    continue
                final_url = normalize_url(response.url) or url
                result.pages.append(PageRecord(final_url, response.status_code, html, True, depth))
                if self._is_contact(final_url):
                    result.contact_page = final_url
                if depth < self.config.max_depth:
                    children = self._links(html, final_url)
                    children.sort(key=lambda item: self._priority(item))
                    queue.extend((child, depth + 1) for child in children)
            except requests.RequestException as exc:
                self._issue(prospect, url, "page_error", "", str(exc), "crawl")

    @staticmethod
    def _priority(url: str) -> int:
        path = urlparse(url).path.lower()
        return 0 if any(term in path for term in PRIORITY_TERMS) else 1

    @staticmethod
    def _is_contact(url: str) -> bool:
        path = urlparse(url).path.lower()
        return any(term in path for term in ("contact", "get-in-touch", "schedule", "appointment"))


# ---------------------------------------------------------------------------
# HTML, JSON-LD, contact, chat and booking extraction
# ---------------------------------------------------------------------------
def valid_email(candidate: str) -> bool:
    """Reject placeholders, tracking/dependency domains and asset-like values."""
    email = candidate.strip().lower().rstrip(".,;:)")
    if not EMAIL_RE.fullmatch(email):
        return False
    local, domain = email.rsplit("@", 1)
    if local in EXCLUDED_EMAIL_LOCAL_PARTS or domain in EXCLUDED_EMAIL_DOMAINS:
        return False
    if domain.rsplit(".", 1)[-1] in {"png", "jpg", "jpeg", "gif", "js", "css"}:
        return False
    if domain.endswith(("sentry.io", "sentry.wixpress.com", "wixpress.com")):
        return False
    if any(token in local for token in ("webpack", "bundle", "tracking", "analytics", "pixel")):
        return False
    if Path(local).suffix.lower() in {".png", ".jpg", ".jpeg", ".gif", ".js"}:
        return False
    return True


def extract_emails(html: str) -> str:
    """Compatibility helper returning cleaned email candidates from raw HTML."""
    candidates = sorted({email.lower() for email in EMAIL_RE.findall(html) if valid_email(email)})
    return "; ".join(candidates) if candidates else "N/A"


def extract_email_evidence(page: PageRecord) -> list[EmailEvidence]:
    """Extract only visible/contact/JSON-LD email evidence from a page."""
    soup = BeautifulSoup(page.html, "html.parser")
    evidence: list[EmailEvidence] = []
    page_path = urlparse(page.url).path.lower()
    role = "contact" if any(term in page_path for term in ("contact", "get-in-touch")) else "other"
    if any(term in page_path for term in ("about", "team")):
        role = "about"

    for script in soup.find_all("script"):
        if (script.get("type") or "").lower() == "application/ld+json":
            for email in EMAIL_RE.findall(script.string or script.get_text()):
                if valid_email(email):
                    evidence.append(EmailEvidence(email.lower(), page.url, "jsonld", "MEDIUM"))

    for anchor in soup.find_all("a", href=True):
        href = anchor.get("href", "")
        if href.lower().startswith("mailto:"):
            email = href[7:].split("?", 1)[0]
            if valid_email(email):
                evidence.append(EmailEvidence(email.lower(), page.url, "mailto", "MEDIUM"))

    for region_name in ("header", "footer"):
        for region in soup.find_all(region_name):
            text = region.get_text(" ", strip=True)
            for email in EMAIL_RE.findall(text):
                if valid_email(email):
                    evidence.append(EmailEvidence(email.lower(), page.url, region_name, "HIGH"))

    visible_soup = BeautifulSoup(page.html, "html.parser")
    for tag in visible_soup.find_all(["script", "style", "noscript", "template"]):
        tag.decompose()
    for email in EMAIL_RE.findall(visible_soup.get_text(" ", strip=True)):
        if valid_email(email):
            confidence = "HIGH" if role == "contact" else "MEDIUM" if role == "about" else "LOW"
            evidence.append(EmailEvidence(email.lower(), page.url, role, confidence))

    # Deduplicate while retaining the strongest source for each address.
    rank = {"contact": 0, "footer": 1, "header": 1, "about": 2, "jsonld": 3, "mailto": 4, "other": 5}
    best: dict[str, EmailEvidence] = {}
    for item in evidence:
        if item.email not in best or rank[item.source_type] < rank[best[item.email].source_type]:
            best[item.email] = item
    return list(best.values())


def jsonld_objects(pages: Iterable[PageRecord]) -> tuple[list[dict[str, Any]], set[str]]:
    """Parse JSON-LD objects and collect schema types without trusting free text."""
    objects: list[dict[str, Any]] = []
    types: set[str] = set()
    for page in pages:
        soup = BeautifulSoup(page.html, "html.parser")
        for script in soup.find_all("script", type=lambda value: value and value.lower() == "application/ld+json"):
            try:
                loaded = json.loads(script.string or script.get_text())
            except (json.JSONDecodeError, TypeError):
                continue
            candidates = loaded if isinstance(loaded, list) else [loaded]
            for item in candidates:
                if not isinstance(item, dict):
                    continue
                if isinstance(item.get("@graph"), list):
                    candidates.extend(obj for obj in item["@graph"] if isinstance(obj, dict))
                if "@type" in item:
                    raw_types = item["@type"] if isinstance(item["@type"], list) else [item["@type"]]
                    types.update(str(value) for value in raw_types)
                objects.append(item)
    return objects, types


def recursive_values(objects: Iterable[dict[str, Any]], key: str) -> list[str]:
    values: list[str] = []
    def visit(value: Any) -> None:
        if isinstance(value, dict):
            if key in value:
                current = value[key]
                if isinstance(current, list):
                    values.extend(safe_text(item) for item in current)
                else:
                    values.append(safe_text(current))
            for child in value.values():
                if isinstance(child, (dict, list)):
                    visit(child)
        elif isinstance(value, list):
            for child in value:
                visit(child)
    for obj in objects:
        visit(obj)
    return list(dict.fromkeys(item for item in values if item not in {"UNKNOWN", ""}))


def detect_provider(source: str, providers: dict[str, tuple[str, ...]]) -> tuple[bool, str, str]:
    lowered = source.lower()
    for provider, markers in providers.items():
        for marker in markers:
            if marker in lowered:
                return True, provider, marker
    return False, "UNKNOWN", ""


def extract_booking(pages: Iterable[PageRecord]) -> tuple[bool, str, str]:
    for page in pages:
        soup = BeautifulSoup(page.html, "html.parser")
        source = page.html.lower()
        for anchor in soup.find_all("a", href=True):
            href = normalize_url(anchor.get("href"), page.url)
            probe = f"{href} {anchor.get_text(' ', strip=True)}".lower()
            found, provider, _ = detect_provider(probe, BOOKING_PROVIDERS)
            if found:
                return True, provider, href or anchor.get("href", "")
        found, provider, _ = detect_provider(source, BOOKING_PROVIDERS)
        if found:
            return True, provider, page.url
    return False, "UNKNOWN", "UNKNOWN"


def extract_chat(pages: Iterable[PageRecord]) -> tuple[bool, str, bool]:
    source = " ".join(page.html.lower() for page in pages)
    has_chat, provider, _ = detect_provider(source, CHAT_PROVIDERS)
    soup_text = " ".join(BeautifulSoup(page.html, "html.parser").get_text(" ", strip=True).lower() for page in pages)
    explicit_ai = any(term in f"{source} {soup_text}" for term in (
        "ai assistant", "ai chatbot", "artificial intelligence", "virtual assistant", "ask our ai",
    ))
    return has_chat, provider, bool(has_chat and explicit_ai)


def _dedupe_text(values: Iterable[str], limit: int = 12) -> str:
    clean = []
    for value in values:
        value = value.strip()
        if value and value not in {"UNKNOWN", "N/A"} and value not in clean:
            clean.append(value)
    return "; ".join(clean[:limit]) if clean else "UNKNOWN"


def extract_site_data(prospect: Prospect, crawl: CrawlResult) -> ExtractedData:
    """Build a factual business profile from crawled pages and schema."""
    data = ExtractedData()
    if not crawl.pages:
        return data
    objects, schema_types = jsonld_objects(crawl.pages)
    data.jsonld_found = bool(objects)
    data.schema_types = _dedupe_text(sorted(schema_types))
    pages = crawl.pages
    emails: list[EmailEvidence] = []
    for page in pages:
        emails.extend(extract_email_evidence(page))
    if emails:
        emails.sort(key=lambda item: ({"HIGH": 0, "MEDIUM": 1, "LOW": 2}[item.confidence], item.source_url))
        data.email_candidates = "; ".join(dict.fromkeys(item.email for item in emails))
        best = emails[0]
        data.verified_email, data.email_confidence, data.email_source_url = best.email, best.confidence, best.source_url

    phones: list[tuple[str, str]] = []
    for page in pages:
        for phone in PHONE_RE.findall(page.html):
            normalized = format_us_phone(phone)
            if normalized != "N/A":
                phones.append((normalized, page.url))
    schema_phones = recursive_values(objects, "telephone")
    for phone in schema_phones:
        normalized = format_us_phone(phone)
        if normalized != "N/A":
            phones.append((normalized, pages[0].url))
    if phones:
        data.verified_phone, data.phone_source_url = phones[0]

    names = recursive_values(objects, "name")
    data.business_name = names[0] if names else prospect.company_name
    services = recursive_values(objects, "service") + recursive_values(objects, "hasOfferCatalog")
    heading_services = []
    for page in pages:
        soup = BeautifulSoup(page.html, "html.parser")
        heading_services.extend(tag.get_text(" ", strip=True) for tag in soup.find_all(["h1", "h2", "h3"]))
    data.services = _dedupe_text(services + [item for item in heading_services if 2 <= len(item.split()) <= 8])
    data.services_source_urls = "; ".join(dict.fromkeys(page.url for page in pages if page.html and data.services != "UNKNOWN")) or "UNKNOWN"
    data.service_area = _dedupe_text(recursive_values(objects, "areaServed"))
    locations = recursive_values(objects, "address") + recursive_values(objects, "location")
    data.locations = _dedupe_text(locations)
    hours = recursive_values(objects, "openingHours") + recursive_values(objects, "openingHoursSpecification")
    data.office_hours = _dedupe_text(hours)
    data.hours_source_url = pages[0].url if data.office_hours != "UNKNOWN" else "UNKNOWN"

    source = " ".join(page.html.lower() for page in pages)
    text = " ".join(BeautifulSoup(page.html, "html.parser").get_text(" ", strip=True).lower() for page in pages)
    data.faq_detected = bool("faq" in source or "frequently asked" in text)
    data.has_chat, data.chat_provider, data.has_ai_chat = extract_chat(pages)
    data.has_booking, data.booking_provider, data.booking_url = extract_booking(pages)
    if data.has_booking:
        data.booking_source_url = next((page.url for page in pages if data.booking_url in page.html or data.booking_url == page.url), pages[0].url)

    for page in pages:
        soup = BeautifulSoup(page.html, "html.parser")
        for form in soup.find_all("form"):
            form_text = f"{form.get_text(' ', strip=True)} {form.get('action', '')}".lower()
            data.has_contact_form |= any(word in form_text for word in ("contact", "message", "name", "email", "phone"))
            data.has_quote_form |= any(word in form_text for word in ("quote", "estimate", "pricing", "service request"))
    data.has_lead_capture = data.has_contact_form or data.has_quote_form or data.has_booking or data.has_chat
    explicit_24_7 = any(term in f"{source} {text}" for term in ("24/7", "24 hours", "open 24 hours", "always available"))
    if explicit_24_7:
        data.limited_business_hours, data.after_hours_gap = "FALSE", "FALSE"
    elif data.office_hours != "UNKNOWN":
        data.limited_business_hours = "TRUE"
        data.after_hours_gap = "TRUE" if not data.has_ai_chat else "FALSE"

    buttons = []
    for page in pages:
        soup = BeautifulSoup(page.html, "html.parser")
        buttons.extend(tag.get_text(" ", strip=True) for tag in soup.find_all(["button", "a"]) if tag.get_text(" ", strip=True))
    cta = next((item for item in buttons if any(term in item.lower() for term in ("quote", "estimate", "schedule", "book", "contact", "consult"))), "")
    data.main_cta = cta or "UNKNOWN"
    lower_cta = cta.lower()
    if any(term in lower_cta for term in ("quote", "estimate")):
        data.main_conversion_goal = "REQUEST_QUOTE"
    elif any(term in lower_cta for term in ("schedule", "book", "appointment")):
        data.main_conversion_goal = "BOOK_APPOINTMENT"
    elif "contact" in lower_cta or data.has_contact_form:
        data.main_conversion_goal = "CONTACT_BUSINESS"
    data.business_model = infer_business_model(prospect)
    data.high_ticket_likelihood = "TRUE" if prospect.target_category in {"Home Services", "Law Firms"} or "Commercial" in prospect.search_query else "UNKNOWN"
    sophistication_signals = sum((data.has_booking, data.has_chat, data.has_contact_form, data.has_quote_form, len(pages) >= 5))
    data.business_sophistication = "HIGH" if sophistication_signals >= 4 else "MEDIUM" if sophistication_signals >= 2 else "LOW"
    return data


def infer_business_model(prospect: Prospect) -> str:
    query = f"{prospect.target_category} {prospect.search_query}".lower()
    if "commercial real estate" in query:
        return "commercial_real_estate"
    if prospect.target_category == "Real Estate":
        return "real_estate"
    if prospect.target_category == "Law Firms":
        return "legal"
    if prospect.target_category == "Home Services":
        return "home_services"
    return "generic_b2b"


# ---------------------------------------------------------------------------
# Deterministic AI Front Desk opportunity analysis
# ---------------------------------------------------------------------------
def bool_value(value: bool | str) -> bool:
    return value is True or str(value).upper() == "TRUE"


def calculate_ai_fit_score(prospect: Prospect, data: ExtractedData, config: Config, crawl: CrawlResult | None = None) -> int:
    """Calculate a bounded 0-100 score from observable signals only."""
    score = 0
    high_ticket = data.high_ticket_likelihood == "TRUE"
    score += config.weights["high_ticket_industry"] if high_ticket else 12 if data.high_ticket_likelihood == "UNKNOWN" else 0
    score += config.weights["no_ai_chat"] if not data.has_ai_chat else 0
    score += config.weights["no_easy_booking"] if not data.has_booking else 0
    score += config.weights["after_hours_gap"] if data.after_hours_gap == "TRUE" else 4 if data.after_hours_gap == "UNKNOWN" else 0
    contact_score = 15 if data.verified_email != "N/A" or data.verified_phone != "N/A" else 8 if prospect.phone_original != "N/A" else 0
    score += min(config.weights["easy_contact"], contact_score)
    evidence_count = sum((data.services != "UNKNOWN", data.locations != "UNKNOWN", data.office_hours != "UNKNOWN", data.faq_detected, data.jsonld_found))
    score += min(config.weights["public_demo_data"], evidence_count * 2)
    strong_signals = prospect.google_review_count >= 20 and isinstance(prospect.google_rating, (int, float)) and float(prospect.google_rating) >= 4.0
    score += config.weights["strong_business_signals"] if strong_signals else 5 if prospect.google_review_count >= 20 else 0

    if data.has_booking:
        score -= config.penalties["effective_booking"]
    if data.has_ai_chat:
        score -= config.penalties["sophisticated_chatbot"]
    if data.has_ai_chat and (data.has_booking or data.after_hours_gap == "FALSE"):
        score -= config.penalties["strong_24_7_intake"]
    if data.business_sophistication == "HIGH":
        score -= config.penalties["sophisticated_funnel"]
    if "health" in prospect.search_query.lower() or "medical" in prospect.search_query.lower():
        score -= config.penalties["healthcare_complexity"]
    if prospect.target_category == "Law Firms":
        score -= config.penalties["legal_complexity"]
    if data.business_sophistication == "HIGH" and crawl and isinstance(crawl.estimated_page_count, int) and crawl.estimated_page_count >= 20:
        score -= config.penalties["large_enterprise"]
    return max(0, min(100, int(score)))


def opportunity_tier(score: int, config: Config) -> str:
    if score >= config.tier_thresholds["A+"]:
        return "A+"
    if score >= config.tier_thresholds["A"]:
        return "A"
    if score >= config.tier_thresholds["B"]:
        return "B"
    if score >= config.tier_thresholds["C"]:
        return "C"
    return "D"


def recommended_action(score: int, config: Config) -> str:
    if score >= 80:
        return "BUILD_DEMO"
    if score >= 70:
        return "CONTACT_DIRECTLY"
    if score >= 40:
        return "REVIEW_MANUALLY"
    return "SKIP"


def first_question(data: ExtractedData, prospect: Prospect) -> str:
    if data.business_model == "home_services":
        return "What type of service do you need, and what is the property ZIP code?"
    if data.business_model in {"real_estate", "commercial_real_estate"}:
        return "Are you looking to buy, sell, lease, or manage a property?"
    if data.business_model == "legal":
        return "What type of legal matter would you like help with?"
    return f"How can {prospect.company_name} help you today?"


def qualification_flow(data: ExtractedData) -> str:
    return data.business_model if data.business_model in {"home_services", "real_estate", "commercial_real_estate", "legal", "generic_b2b"} else "manual_review"


def build_sales_angle(prospect: Prospect, data: ExtractedData, crawl: CrawlResult) -> tuple[str, str, str]:
    evidence = []
    urls = []
    if data.has_contact_form and not data.has_chat:
        evidence.append("The website captures inquiries, but no conversational qualification was detected before the form.")
        urls.append(data.contact_page if data.contact_page != "UNKNOWN" else crawl.final_url)
    elif data.has_booking:
        evidence.append("The company already exposes booking, so the opportunity is better suited to qualification before scheduling.")
        urls.append(data.booking_url if data.booking_url != "UNKNOWN" else data.booking_source_url)
    elif data.after_hours_gap == "TRUE":
        evidence.append("Published office hours are limited and no automated after-hours intake was detected.")
        urls.append(data.hours_source_url)
    elif data.services != "UNKNOWN":
        evidence.append("Services are publicly identifiable, creating a usable basis for a tailored intake and qualification flow.")
        urls.append(data.services_source_urls)
    else:
        evidence.append("The public site has limited conversion evidence; a manual review is recommended before outreach.")
        urls.append(crawl.final_url)
    angle = evidence[0]
    evidence_text = "Observable evidence: " + angle
    evidence_urls = "; ".join(item for item in urls if item and item != "UNKNOWN") or "UNKNOWN"
    return angle, evidence_text, evidence_urls


def apply_analysis(prospect: Prospect, crawl: CrawlResult, data: ExtractedData, config: Config) -> None:
    """Copy validation/profile observations and score into the flat prospect."""
    for key, value in asdict(data).items():
        if hasattr(prospect, key):
            setattr(prospect, key, value)
    prospect.verified_website = crawl.final_url if crawl.website_reachable else "N/A"
    prospect.website_reachable = crawl.website_reachable
    prospect.final_url = crawl.final_url or "UNKNOWN"
    prospect.final_http_status = crawl.final_http_status or "UNKNOWN"
    prospect.redirect_count = crawl.redirect_count
    prospect.homepage_status = crawl.homepage_status or "UNKNOWN"
    prospect.internal_link_count = crawl.internal_link_count
    prospect.estimated_page_count = crawl.estimated_page_count
    prospect.sitemap_found = crawl.sitemap_found
    prospect.robots_txt_found = crawl.robots_txt_found
    prospect.javascript_heavy = crawl.javascript_heavy
    prospect.crawl_status = crawl.crawl_status
    prospect.crawl_error = crawl.crawl_error
    prospect.contact_page = crawl.contact_page
    if prospect.validation_status == "UNVERIFIED" and crawl.website_reachable:
        prospect.validation_status = "VERIFIED"
    prospect.phone = data.verified_phone if data.verified_phone != "N/A" else prospect.phone_original
    prospect.email = data.verified_email
    prospect.ai_fit_score = calculate_ai_fit_score(prospect, data, config, crawl)
    prospect.opportunity_tier = opportunity_tier(prospect.ai_fit_score, config)
    prospect.recommended_action = recommended_action(prospect.ai_fit_score, config)
    prospect.qualification_flow_recommendation = qualification_flow(data)
    prospect.suggested_first_question = first_question(data, prospect)
    prospect.suggested_conversion_goal = data.main_conversion_goal if data.main_conversion_goal != "UNKNOWN" else "CAPTURE_AND_QUALIFY_LEAD"
    prospect.sales_angle, prospect.sales_angle_evidence, prospect.evidence_urls = build_sales_angle(prospect, data, crawl)


# ---------------------------------------------------------------------------
# Timestamped professional Excel export
# ---------------------------------------------------------------------------
def unique_output_path(output_dir: Path, run_time: datetime | None = None) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = (run_time or datetime.now().astimezone()).strftime("%Y-%m-%d_%H-%M-%S")
    path = output_dir / f"form4th_prospects_{stamp}.xlsx"
    suffix = 2
    while path.exists():
        path = output_dir / f"form4th_prospects_{stamp}_{suffix:02d}.xlsx"
        suffix += 1
    return path


PROSPECT_COLUMNS = list(asdict(Prospect()).keys())
TOP_COLUMNS = [
    "company_name", "target_category", "city", "state", "verified_website", "verified_email",
    "verified_phone", "ai_fit_score", "opportunity_tier", "has_chat", "has_ai_chat",
    "has_booking", "after_hours_gap", "high_ticket_likelihood", "sales_angle", "recommended_action",
]
DEMO_COLUMNS = TOP_COLUMNS + [
    "services", "qualification_flow_recommendation", "suggested_first_question",
    "suggested_conversion_goal", "evidence_urls",
]


def _style_sheet(worksheet: Any, centered: set[str] | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF")
    body_font = Font(name="Segoe UI", size=10, color="111827")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centered = centered or set()
    header_indexes = {cell.value: cell.column for cell in worksheet[1]}
    for cell in worksheet[1]:
        cell.fill, cell.font, cell.border = header_fill, header_font, border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 32
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font, cell.border = body_font, border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for name in centered:
            if name in header_indexes:
                worksheet.cell(row[0].row, header_indexes[name]).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max(length + 2, 12), 55)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def export_workbook(prospects: list[Prospect], issues: list[ValidationIssue], run_summary: dict[str, Any], output_path: Path) -> Path:
    """Write all required sheets without overwriting a previous execution."""
    dataframe = pd.DataFrame([item.to_dict() for item in prospects], columns=PROSPECT_COLUMNS)
    dataframe.sort_values(by=["ai_fit_score", "google_review_count"], ascending=[False, False], inplace=True, ignore_index=True)
    top = dataframe[dataframe["ai_fit_score"] >= 70][TOP_COLUMNS].copy()
    demos = dataframe[dataframe["ai_fit_score"] >= 80][DEMO_COLUMNS].copy()
    issues_df = pd.DataFrame([asdict(item) for item in issues], columns=list(asdict(ValidationIssue("", "", "", "", "", "", "")).keys()))
    summary_df = pd.DataFrame([run_summary])
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Prospects")
        top.to_excel(writer, index=False, sheet_name="Top Opportunities")
        demos.to_excel(writer, index=False, sheet_name="Demo Candidates")
        issues_df.to_excel(writer, index=False, sheet_name="Validation Issues")
        summary_df.to_excel(writer, index=False, sheet_name="Run Summary")

    workbook = load_workbook(output_path)
    for sheet_name in workbook.sheetnames:
        centered = {"validation_status", "opportunity_tier", "recommended_action", "ai_fit_score", "has_chat", "has_ai_chat", "has_booking", "website_reachable"}
        if sheet_name == "Run Summary":
            centered = set()
        _style_sheet(workbook[sheet_name], centered)
    prospects_sheet = workbook["Prospects"]
    headers = {cell.value: cell.column for cell in prospects_sheet[1]}
    tier_col = prospects_sheet.cell(1, headers["opportunity_tier"]).column_letter
    colors = {"A+": "63BE7B", "A": "A9D18E", "B": "FFEB9C", "C": "F4B183", "D": "F8696B"}
    for tier, color in colors.items():
        fill = PatternFill("solid", fgColor=color)
        for row in range(2, prospects_sheet.max_row + 1):
            cell = prospects_sheet[f"{tier_col}{row}"]
            if cell.value == tier:
                cell.fill = fill
                cell.font = Font(name="Segoe UI", bold=True, color="111827")
    for worksheet in workbook.worksheets:
        worksheet.sheet_view.showGridLines = False
    workbook.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# CLI and orchestration
# ---------------------------------------------------------------------------
def load_query_config(path: Path) -> dict[str, list[str]]:
    """Load a small JSON search config; YAML can be added without changing the engine."""
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise SystemExit(f"Config must be valid JSON for now: {path}") from exc
    queries = payload.get("queries", payload)
    if not isinstance(queries, dict):
        raise SystemExit("Config must contain an object named 'queries'.")
    return {str(category): [str(query) for query in values] for category, values in queries.items() if isinstance(values, list)}


def query_plan(args: argparse.Namespace) -> dict[str, list[str]]:
    if args.config:
        return load_query_config(Path(args.config))
    if not args.category and not args.city:
        return DEFAULT_QUERIES
    category = CATEGORY_ALIASES.get((args.category or "").lower(), args.category) or "Home Services"
    city = args.city or "Austin, TX"
    if category == "Home Services":
        templates = ["Roofing contractors in {city}", "HVAC services in {city}", "Plumbers in {city}"]
    elif category == "Law Firms":
        templates = ["Personal injury lawyer in {city}", "Family law attorney in {city}", "Criminal defense lawyer in {city}"]
    else:
        templates = ["Property management in {city}", "Commercial real estate in {city}"]
    return {category: [template.format(city=city) for template in templates]}


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--api-key", default="", help="Google Places key; env GOOGLE_PLACES_API_KEY is preferred.")
    parser.add_argument("--category", help="Category: Home Services/home_services, Law Firms/law_firms or Real Estate/real_estate.")
    parser.add_argument("--city", help='City/state, e.g. "Austin, TX".')
    parser.add_argument("--config", help="JSON file containing {\"queries\": {\"Category\": [\"query\"]}}.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--max-pages", type=int, default=12)
    parser.add_argument("--max-depth", type=int, default=1)
    parser.add_argument("--timeout", type=int, default=5)
    parser.add_argument("--min-reviews", type=int, default=0)
    return parser.parse_args(argv)


def run_engine(args: argparse.Namespace, discovery: PlacesNewDiscovery | None = None) -> Path:
    start = datetime.now().astimezone()
    run_id = start.strftime("%Y%m%d%H%M%S")
    config = Config(
        max_pages_per_site=max(1, args.max_pages),
        max_depth=max(0, args.max_depth),
        request_timeout=max(1, args.timeout),
        min_reviews=max(0, args.min_reviews),
        output_dir=Path(args.output_dir),
    )
    queries_by_category = query_plan(args)
    api_key = (args.api_key or os.getenv("GOOGLE_PLACES_API_KEY") or GOOGLE_PLACES_API_KEY).strip()
    if not api_key:
        raise SystemExit("ERROR: set GOOGLE_PLACES_API_KEY or provide --api-key.")

    discovery = discovery or PlacesNewDiscovery(api_key, config)
    raw_prospects: list[Prospect] = []
    total_queries = sum(len(items) for items in queries_by_category.values())
    completed_queries = 0
    for category, queries in queries_by_category.items():
        for query in queries:
            completed_queries += 1
            print(f"[DISCOVERY] ({completed_queries}/{total_queries}) {query}")
            try:
                places = discovery.search(query)
                raw_prospects.extend(build_place_prospect(place, category, query) for place in places)
                print(f"[DISCOVERY] Found {len(places)} businesses")
            except Exception as exc:
                print(f"[DISCOVERY] ERROR {query}: {exc}")

    prospects = deduplicate_prospects(raw_prospects)
    prospects = [item for item in prospects if item.google_review_count >= config.min_reviews]
    issues: list[ValidationIssue] = []
    crawler = WebCrawler(config, issues)
    for index, prospect in enumerate(prospects, 1):
        print(f"[VALIDATION] {index}/{len(prospects)} {prospect.company_name}")
        crawl = crawler.validate_and_crawl(prospect)
        print(f"[CRAWL] {crawl.final_url or prospect.website_original} - {len(crawl.pages)} pages")
        data = extract_site_data(prospect, crawl)
        apply_analysis(prospect, crawl, data, config)
        print(f"[ANALYSIS] Chat: {'YES' if data.has_chat else 'NO'} | Booking: {'YES' if data.has_booking else 'NO'} | Score: {prospect.ai_fit_score}")

    finish = datetime.now().astimezone()
    counts = {tier: sum(item.opportunity_tier == tier for item in prospects) for tier in ("A+", "A", "B", "C", "D")}
    run_summary = {
        "run_id": run_id,
        "start_timestamp": start.isoformat(timespec="seconds"),
        "finish_timestamp": finish.isoformat(timespec="seconds"),
        "search_queries": " | ".join(query for queries in queries_by_category.values() for query in queries),
        "categories": "; ".join(queries_by_category),
        "total_discovered": len(raw_prospects),
        "total_after_deduplication": len(prospects),
        "total_validated": sum(item.website_reachable for item in prospects),
        "total_invalid": sum(item.validation_status == "INVALID" for item in prospects),
        "total_errors": len(issues),
        "A+": counts["A+"], "A": counts["A"], "B": counts["B"], "C": counts["C"], "D": counts["D"],
        "demo_candidates": sum(item.ai_fit_score >= 80 for item in prospects),
        "outreach_candidates": sum(item.ai_fit_score >= 70 for item in prospects),
        "generated_file_path": "pending",
    }
    output_path = unique_output_path(config.output_dir, start)
    run_summary["generated_file_path"] = str(output_path)
    export_workbook(prospects, issues, run_summary, output_path)
    print(f"[EXPORT] {output_path}")
    return output_path


def main(argv: list[str] | None = None) -> None:
    run_engine(parse_args(argv))


if __name__ == "__main__":
    main()
